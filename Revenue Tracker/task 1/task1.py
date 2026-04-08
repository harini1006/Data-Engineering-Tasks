import pandas as pd
import os
os.chdir(r'C:\Users\Harini.Velmurugan\Documents\task 1')
all_sheets = pd.read_excel('Delta3_Apr.xlsx', sheet_name=None, header=None)
def to_number(val):
    if isinstance(val, str):
        return float(val.replace('£', '').replace(',', ''))
    return float(val)
def extract_sheet(df):
    def get(label):
        return to_number(df[df[0] == label][1].values[0])
   
    core        = get('Salary - Core employees')
    tl          = get('Salary - TL / Managers')
    consultants = get('Salary - Consultants')
    incentive   = get('Performance payments - Incentive & Others')
   
    return {
        'revenue':          get('Revenue'),
        'revenue_pct':      get('Revenue %'),
        'total_workforce':  core + tl + consultants + incentive,
        'tech_salary_pct':  (core + tl + consultants + incentive) / get('Revenue'),
        'total_salary':     get('Total salary allocation for project'),
        'total_salary_pct': get('Total salary allocation %'),
    }
 
results = {}
for sheet_name, df in all_sheets.items():
    results[sheet_name] = extract_sheet(df)
    print(sheet_name, results[sheet_name])
 
from openpyxl import load_workbook
 
wb = load_workbook('Delta3_Output.xlsx')
ws = wb['Revenue ']
 
col_map = {
    'MI - Apr @ 108 ': 4,
    'May':             5,
    'June':            6,
}
 
ROW_REVENUE     = 7
ROW_REVENUE_PCT = 8
ROW_WORKFORCE   = 10
ROW_TECH_PCT    = 11
ROW_SALARY      = 13
ROW_SALARY_PCT  = 14
 
for sheet_name, data in results.items():
    col = col_map[sheet_name]
    ws.cell(row=ROW_REVENUE,     column=col).value = data['revenue']
    ws.cell(row=ROW_REVENUE_PCT, column=col).value = data['revenue_pct']
    ws.cell(row=ROW_WORKFORCE,   column=col).value = data['total_workforce']
    ws.cell(row=ROW_TECH_PCT, column=col).value = data['tech_salary_pct']
    ws.cell(row=ROW_SALARY,      column=col).value = data['total_salary']
    ws.cell(row=ROW_SALARY_PCT,  column=col).value = data['total_salary_pct']
 
wb.save('Delta3_Ouptut.xlsx')
print('Done!')
 